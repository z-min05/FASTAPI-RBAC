import pytest
from httpx import AsyncClient


async def _create_project(client: AsyncClient, headers: dict, code: str = "TC-PROJ") -> int:
    resp = await client.post("/api/v1/projects", headers=headers, json={
        "code": code,
        "name": "用例测试项目",
        "is_active": True,
    })
    assert resp.status_code == 200
    return resp.json()["data"]["id"]


@pytest.mark.asyncio
async def test_testcases_unauthorized(client: AsyncClient):
    """未登录访问用例接口应返回 401"""
    response = await client.get("/api/v1/testcases")
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_testcase_crud(client: AsyncClient, superuser_headers: dict):
    """超管：用例新增/列表/详情/编辑/删除/批量删除全流程"""
    project_id = await _create_project(client, superuser_headers, "TC-CRUD")

    # 创建
    resp = await client.post("/api/v1/testcases", headers=superuser_headers, json={
        "project_id": project_id,
        "title": "登录成功",
        "module": "auth",
        "priority": "P0",
        "case_type": "interface",
        "source": "接口文档",
        "precondition": "用户已注册",
        "steps": "1.调用 /auth/login\n2.校验返回 token",
        "expected_result": "返回 200 与 access_token",
    })
    assert resp.status_code == 200
    data = resp.json()["data"]
    testcase_id = data["id"]
    assert data["project_code"] == "TC-CRUD"  # 列表/详情应回显项目编码
    assert data["priority"] == "P0"

    # 列表
    resp = await client.get("/api/v1/testcases", headers=superuser_headers, params={"project_id": project_id})
    assert resp.status_code == 200
    assert resp.json()["data"]["total"] == 1

    # 模块列表
    resp = await client.get("/api/v1/testcases/modules", headers=superuser_headers, params={"project_id": project_id})
    assert resp.status_code == 200
    assert "auth" in resp.json()["data"]

    # 详情
    resp = await client.get(f"/api/v1/testcases/{testcase_id}", headers=superuser_headers)
    assert resp.status_code == 200
    assert resp.json()["data"]["title"] == "登录成功"

    # 编辑
    resp = await client.put(f"/api/v1/testcases/{testcase_id}", headers=superuser_headers, json={
        "priority": "P1",
        "status": "reviewed",
    })
    assert resp.status_code == 200
    assert resp.json()["data"]["priority"] == "P1"
    assert resp.json()["data"]["status"] == "reviewed"

    # 单条删除
    resp = await client.delete(f"/api/v1/testcases/{testcase_id}", headers=superuser_headers)
    assert resp.status_code == 200

    # 批量删除
    ids = []
    for i in range(3):
        resp = await client.post("/api/v1/testcases", headers=superuser_headers, json={
            "project_id": project_id,
            "title": f"批量用例-{i}",
            "module": "batch",
            "expected_result": "成功",
        })
        ids.append(resp.json()["data"]["id"])
    resp = await client.post("/api/v1/testcases/batch-delete", headers=superuser_headers, json={"ids": ids})
    assert resp.status_code == 200
    resp = await client.get("/api/v1/testcases", headers=superuser_headers)
    assert resp.json()["data"]["total"] == 0


@pytest.mark.asyncio
async def test_testcase_invalid_project(client: AsyncClient, superuser_headers: dict):
    """所属项目不存在时应返回 400"""
    resp = await client.post("/api/v1/testcases", headers=superuser_headers, json={
        "project_id": 99999,
        "title": "无效项目",
        "module": "auth",
        "expected_result": "成功",
    })
    assert resp.status_code == 400


@pytest.mark.asyncio
async def test_testcase_inactive_project(client: AsyncClient, superuser_headers: dict):
    """项目停用后不允许新增用例"""
    project_id = await _create_project(client, superuser_headers, "TC-INACTIVE")
    resp = await client.put(f"/api/v1/projects/{project_id}", headers=superuser_headers, json={"is_active": False})
    assert resp.status_code == 200

    resp = await client.post("/api/v1/testcases", headers=superuser_headers, json={
        "project_id": project_id,
        "title": "停用项目下新增",
        "module": "auth",
        "expected_result": "成功",
    })
    assert resp.status_code == 400


@pytest.mark.asyncio
async def test_testcase_export_and_import(client: AsyncClient, superuser_headers: dict):
    """导出 CSV 内容正确；导入 CSV 成功写入"""
    project_id = await _create_project(client, superuser_headers, "TC-EXIM")
    await client.post("/api/v1/testcases", headers=superuser_headers, json={
        "project_id": project_id,
        "title": "导出用例",
        "module": "exim",
        "priority": "P2",
        "expected_result": "成功",
    })

    # 导出
    resp = await client.post("/api/v1/testcases/export", headers=superuser_headers, params={"project_id": project_id})
    assert resp.status_code == 200
    content = resp.json()["data"]["content"]
    assert content.startswith("\ufeff")  # UTF-8 BOM
    assert "导出用例" in content
    assert "TC-EXIM" in content

    # 导入（注释行 + 正确表头 + 两行数据；注释行应被忽略）
    csv_content = (
        "\ufeff# 必填列：项目编码、标题、模块、预期结果；其余列选填\n"
        "项目编码,标题,模块,优先级,类型,来源,前置条件,步骤,预期结果,状态,标签\n"
        "TC-EXIM,导入用例A,moduleA,P1,function,需求文档,无,1.操作,成功,draft,tagA\n"
        "TC-EXIM,导入用例B,moduleB,P3,interface,接口文档,无,1.操作,成功,reviewed,tagB\n"
    )
    resp = await client.post("/api/v1/testcases/import", headers=superuser_headers, json={"content": csv_content})
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["success"] == 2
    assert data["failures"] == []

    resp = await client.get("/api/v1/testcases", headers=superuser_headers, params={"project_id": project_id})
    assert resp.json()["data"]["total"] == 3


@pytest.mark.asyncio
async def test_testcase_import_template(client: AsyncClient, superuser_headers: dict):
    """导入模板下载：xlsx，表头标注（必填）/（非必填）、加粗、"必填"红色"""
    import base64
    import io
    from openpyxl import load_workbook

    resp = await client.get("/api/v1/testcases/import-template", headers=superuser_headers)
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["filename"].endswith(".xlsx")
    assert data["is_base64"] is True

    wb = load_workbook(io.BytesIO(base64.b64decode(data["content"])), rich_text=True)
    ws = wb.active
    headers = [str(c) for c in next(ws.iter_rows(values_only=True))]
    assert "项目编码（必填）" in headers[0]
    assert "标题（必填）" in headers[1]
    assert "优先级（非必填）" in headers[3]
    # 表头加粗且"必填"红色（富文本内联样式）
    rich_cell = ws.cell(row=1, column=1)
    rich_text = rich_cell.value
    red_bold = any(getattr(b, "font", None) and b.font.color and b.font.color.rgb == "FFFF0000" and b.font.b for b in rich_text)
    assert red_bold
    bold_parts = [b for b in rich_text if getattr(b, "font", None) and b.font.b]
    assert len(bold_parts) == len([t for t in rich_text])
    # 示例行
    assert "示例用例" in str(ws.cell(row=2, column=2).value)


@pytest.mark.asyncio
async def test_testcase_import_xlsx(client: AsyncClient, superuser_headers: dict):
    """导入 xlsx 模板文件成功写入"""
    import base64
    import io
    from openpyxl import Workbook

    project_id = await _create_project(client, superuser_headers, "TC-XLSX")

    wb = Workbook()
    ws = wb.active
    ws.append(["项目编码", "标题", "模块", "优先级", "类型", "来源",
               "前置条件", "步骤", "预期结果", "状态", "标签"])
    ws.append(["TC-XLSX", "Excel导入用例A", "moduleX", "P1", "function", "需求文档",
               "无", "1.操作", "成功", "draft", "tagX"])
    ws.append(["TC-XLSX", "Excel导入用例B", "moduleY", "P3", "interface", "接口文档",
               "无", "1.操作", "成功", "reviewed", "tagY"])
    buf = io.BytesIO()
    wb.save(buf)

    resp = await client.post("/api/v1/testcases/import", headers=superuser_headers, json={
        "content": base64.b64encode(buf.getvalue()).decode("ascii"),
        "format": "xlsx",
    })
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["success"] == 2
    assert data["failures"] == []

    resp = await client.get("/api/v1/testcases", headers=superuser_headers, params={"project_id": project_id})
    assert resp.json()["data"]["total"] == 2


@pytest.mark.asyncio
async def test_testcase_import_validation(client: AsyncClient, superuser_headers: dict):
    """导入校验：表头错误 / 非法行应给出失败明细"""
    # 表头错误
    resp = await client.post("/api/v1/testcases/import", headers=superuser_headers, json={
        "content": "编码,标题\nA,用例\n",
    })
    assert resp.status_code == 400

    # 非法行（项目不存在 / 优先级不合法 / 必填缺失）
    csv_content = (
        "\ufeff项目编码,标题,模块,优先级,类型,来源,前置条件,步骤,预期结果,状态,标签\n"
        "NO-SUCH,用例A,moduleA,P1,function,需求文档,无,1.操作,成功,draft,tagA\n"
        "NO-SUCH,,moduleB,P9,function,接口文档,无,1.操作,,draft,tagB\n"
    )
    resp = await client.post("/api/v1/testcases/import", headers=superuser_headers, json={"content": csv_content})
    assert resp.status_code == 200
    data = resp.json()["data"]
    assert data["success"] == 0
    assert len(data["failures"]) == 2
    assert any("项目编码" in "".join(f["errors"]) for f in data["failures"])
    assert any("优先级不合法" in "".join(f["errors"]) for f in data["failures"])

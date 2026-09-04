import csv
import io

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from sqlalchemy import select, or_, and_
from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.testcase_repo import TestCaseRepository
from app.repositories.project_repo import ProjectRepository
from app.models.testcase import TestCase
from app.models.project import Project
from app.schemas.testcase import (
    TestCaseCreate,
    TestCaseUpdate,
    TestCaseResponse,
    ALLOWED_PRIORITIES,
    ALLOWED_STATUS,
    ALLOWED_CASE_TYPES,
)
from app.core.pagination import PaginationParams, PaginatedResponse
from app.exceptions import NotFoundException, BadRequestException, ConflictException
from app.services.auto_file_service import (
    validate_codes,
    validate_root_path,
    generate_automation_file,
)

# CSV 列定义（导入导出共用）
CSV_COLUMNS = [
    "项目编码", "标题", "模块", "优先级", "类型", "来源",
    "前置条件", "步骤", "预期结果", "状态", "标签",
    "模块编码", "用例编码",
]

# 必填列
REQUIRED_COLUMNS = {"项目编码", "标题", "模块", "预期结果"}

# 表头中的必填/非必填标注后缀
_REQUIRED_SUFFIX = "（必填）"
_OPTIONAL_SUFFIX = "（非必填）"


def _normalize_header(name: str) -> str:
    """去掉表头中的（必填）/（非必填）标注，得到标准列名"""
    name = name.strip()
    for suffix in (_REQUIRED_SUFFIX, _OPTIONAL_SUFFIX):
        if name.endswith(suffix):
            return name[: -len(suffix)].strip()
    return name


class TestCaseService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.testcase_repo = TestCaseRepository(db)
        self.project_repo = ProjectRepository(db)

    # ---------- 查询 ----------

    def _build_filters(
        self,
        project_id: int | None = None,
        module: str | None = None,
        priority: str | None = None,
        status: str | None = None,
        source: str | None = None,
        keyword: str | None = None,
    ) -> list:
        filters = []
        if project_id is not None:
            filters.append(TestCase.project_id == project_id)
        if module:
            filters.append(TestCase.module == module)
        if priority:
            filters.append(TestCase.priority == priority)
        if status:
            filters.append(TestCase.status == status)
        if source:
            filters.append(TestCase.source == source)
        if keyword:
            filters.append(
                or_(
                    TestCase.title.ilike(f"%{keyword}%"),
                    TestCase.module.ilike(f"%{keyword}%"),
                )
            )
        return filters

    async def get_testcase(self, testcase_id: int) -> TestCase:
        tc = await self.testcase_repo.get_by_id(testcase_id)
        if not tc:
            raise NotFoundException("用例不存在")
        return tc

    async def get_testcases(
        self,
        params: PaginationParams,
        project_id: int | None = None,
        module: str | None = None,
        priority: str | None = None,
        status: str | None = None,
        source: str | None = None,
        keyword: str | None = None,
        order: str = "desc",
    ) -> PaginatedResponse:
        filters = self._build_filters(project_id, module, priority, status, source, keyword)
        result = await self.testcase_repo.get_paginated(params, filters or None, order)

        project_map = await self._get_project_map([tc.project_id for tc in result.items])
        items = [self._to_response(tc, project_map) for tc in result.items]
        return PaginatedResponse(
            items=items,
            total=result.total,
            page=result.page,
            page_size=result.page_size,
            total_pages=result.total_pages,
        )

    async def _get_project_map(self, project_ids: list[int]) -> dict[int, Project]:
        ids = list(set(project_ids))
        if not ids:
            return {}
        stmt = select(Project).where(Project.id.in_(ids))
        result = await self.db.execute(stmt)
        return {p.id: p for p in result.scalars().all()}

    def _to_response(self, tc: TestCase, project_map: dict[int, Project]) -> dict:
        proj = project_map.get(tc.project_id)
        return TestCaseResponse(
            id=tc.id,
            project_id=tc.project_id,
            project_code=proj.code if proj else None,
            project_name=proj.name if proj else None,
            title=tc.title,
            module=tc.module,
            priority=tc.priority,
            case_type=tc.case_type,
            source=tc.source,
            precondition=tc.precondition,
            steps=tc.steps,
            expected_result=tc.expected_result,
            status=tc.status,
            tags=tc.tags,
            module_code=tc.module_code,
            case_code=tc.case_code,
            created_at=tc.created_at,
            updated_at=tc.updated_at,
        ).model_dump()

    async def get_modules(self, project_id: int | None = None) -> list[str]:
        return await self.testcase_repo.get_modules(project_id)

    # ---------- 写操作 ----------

    async def create_testcase(self, data: TestCaseCreate) -> TestCase:
        project = await self._ensure_project_active(data.project_id)
        # 空字符串转为 None，避免与数据库条件唯一索引冲突
        module_code = data.module_code or None
        case_code = data.case_code or None
        validate_codes(module_code, case_code)
        await self._check_uniqueness(data.project_id, module_code, case_code)
        raw = data.model_dump()
        raw["module_code"] = module_code
        raw["case_code"] = case_code
        tc = TestCase(**raw)
        created = await self.testcase_repo.create(tc)
        # 尝试生成自动化文件
        await self._generate_auto_file(project, created)
        return created

    async def update_testcase(self, testcase_id: int, data: TestCaseUpdate) -> TestCase:
        tc = await self.testcase_repo.get_by_id(testcase_id)
        if not tc:
            raise NotFoundException("用例不存在")
        update_data = data.model_dump(exclude_unset=True)
        new_project_id = update_data.get("project_id")
        # 空字符串转为 None，避免条件唯一索引冲突
        new_module_code = update_data.pop("module_code", tc.module_code)
        new_module_code = new_module_code or None
        new_case_code = update_data.pop("case_code", tc.case_code)
        new_case_code = new_case_code or None
        update_data["module_code"] = new_module_code
        update_data["case_code"] = new_case_code
        project = None
        if new_project_id is not None and new_project_id != tc.project_id:
            project = await self._ensure_project_active(new_project_id)
        else:
            project = await self.project_repo.get_by_id(new_project_id or tc.project_id)
        # 校验格式与唯一性
        if "module_code" in update_data or "case_code" in update_data:
            validate_codes(new_module_code, new_case_code)
            await self._check_uniqueness(
                new_project_id or tc.project_id,
                new_module_code,
                new_case_code,
                exclude_id=tc.id,
            )
        updated = await self.testcase_repo.update(testcase_id, update_data)
        # 尝试生成自动化文件（用更新后的值）
        if project and updated.module_code and updated.case_code:
            await self._generate_auto_file(project, updated)
        return updated

    async def _ensure_project_active(self, project_id: int) -> Project:
        project = await self.project_repo.get_by_id(project_id)
        if not project:
            raise BadRequestException("项目不存在")
        if not project.is_active:
            raise BadRequestException("项目已停用，不能在该项目下操作用例")
        return project

    async def _check_uniqueness(self, project_id: int, module_code: str | None, case_code: str | None, exclude_id: int | None = None) -> None:
        """校验同一项目下 (module_code, case_code) 组合唯一"""
        if not module_code or not case_code:
            return
        filters = [
            TestCase.project_id == project_id,
            TestCase.module_code == module_code,
            TestCase.case_code == case_code,
        ]
        if exclude_id is not None:
            filters.append(TestCase.id != exclude_id)
        stmt = select(TestCase.id).where(and_(*filters)).limit(1)
        result = (await self.db.execute(stmt)).scalar()
        if result is not None:
            raise ConflictException(
                f"该项目下已存在相同模块编码+用例编码的组合（模块编码: {module_code}，用例编码: {case_code}）"
            )

    async def _generate_auto_file(self, project: Project, tc: TestCase) -> dict:
        """尝试生成自动化文件，返回 {generated: bool,  message: str}"""
        if not (project.auto_root_path and tc.module_code and tc.case_code):
            return {"generated": False, "message": ""}
        ok, msg = generate_automation_file(project.auto_root_path, tc)
        if not ok:
            return {"generated": False, "message": f"自动化文件生成失败: {msg}"}
        return {"generated": True, "message": msg}

    async def delete_testcase(self, testcase_id: int) -> None:
        refs = await self.plan_tc_repo.count_by_testcase(testcase_id)
        if refs > 0:
            raise ConflictException(f"用例已被 {refs} 个测试计划引用，请先从计划中移除后再删除")
        if not await self.testcase_repo.delete(testcase_id):
            raise NotFoundException("用例不存在")

    async def delete_testcases(self, ids: list[int]) -> int:
        refs = await self.plan_tc_repo.count_by_testcases(ids)
        if refs > 0:
            raise ConflictException("存在被测试计划引用的用例，请先从计划中移除后再批量删除")
        return await self.testcase_repo.delete_batch(ids)

    # ---------- CSV 导入导出 ----------

    async def get_import_template(self) -> bytes:
        """导入模板（xlsx）：表头标注（必填）/（非必填），表头加粗，"必填"红色"""
        wb = Workbook()
        ws = wb.active
        ws.title = "用例导入模板"

        def header_cell(name: str) -> CellRichText:
            parts = [TextBlock(InlineFont(b=True), name)]
            if name in REQUIRED_COLUMNS:
                parts.append(TextBlock(InlineFont(b=True), "（"))
                parts.append(TextBlock(InlineFont(b=True, color="FFFF0000"), "必填"))
                parts.append(TextBlock(InlineFont(b=True), "）"))
            else:
                parts.append(TextBlock(InlineFont(b=True), "（非必填）"))
            return CellRichText(parts)

        for i, col in enumerate(CSV_COLUMNS, start=1):
            ws.cell(row=1, column=i, value=header_cell(col))

        # 示例行（按需替换，导入前可删除）
        ws.append([
            "DEMO", "示例用例-登录功能", "login", "P1", "function", "需求文档",
            "已注册测试账号", "1. 打开登录页\n2. 输入账号密码\n3. 点击登录",
            "登录成功并跳转首页", "reviewed", "冒烟,登录",
            "test_login", "test_login_success",
        ])

        for col, width in {"A": 12, "B": 20, "C": 12, "D": 10, "E": 10, "F": 12,
                           "G": 14, "H": 30, "I": 22, "J": 10, "K": 14, "L": 16, "M": 18}.items():
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def export_csv(
        self,
        project_id: int | None = None,
        module: str | None = None,
        priority: str | None = None,
        status: str | None = None,
        source: str | None = None,
        keyword: str | None = None,
    ) -> str:
        """按筛选条件导出全部用例为 CSV 文本（UTF-8 BOM）"""
        filters = self._build_filters(project_id, module, priority, status, source, keyword)
        stmt = select(TestCase).order_by(TestCase.id)
        if filters:
            stmt = stmt.where(*filters)
        result = await self.db.execute(stmt)
        testcases = list(result.scalars().all())
        project_map = await self._get_project_map([tc.project_id for tc in testcases])

        buf = io.StringIO()
        buf.write("\ufeff")
        writer = csv.writer(buf)
        writer.writerow(CSV_COLUMNS)
        for tc in testcases:
            proj = project_map.get(tc.project_id)
            writer.writerow([
                proj.code if proj else "",
                tc.title,
                tc.module,
                tc.priority,
                tc.case_type,
                tc.source or "",
                tc.precondition or "",
                tc.steps or "",
                tc.expected_result,
                tc.status,
                tc.tags or "",
                tc.module_code or "",
                tc.case_code or "",
            ])
        return buf.getvalue()

    async def import_csv(self, content: str) -> dict:
        """导入 CSV，返回 {success, failures:[{line, errors}]}"""
        content = content.lstrip("\ufeff").strip()
        if not content:
            raise BadRequestException("导入内容为空")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        # 跳过以 # 开头的注释行（模板说明行）
        rows = [r for r in rows if not (r and str(r[0]).strip().startswith("#"))]
        if len(rows) < 2:
            raise BadRequestException("CSV 缺少数据行")

        header = [_normalize_header(h) for h in rows[0]]
        if header != CSV_COLUMNS:
            raise BadRequestException("CSV 表头不正确，应为: " + ",".join(CSV_COLUMNS))

        records = []
        for row in rows[1:]:
            if not row or not any(str(c).strip() for c in row):
                continue
            padded = [str(c).strip() for c in row] + [""] * (len(CSV_COLUMNS) - len(row))
            records.append(dict(zip(CSV_COLUMNS, padded[: len(CSV_COLUMNS)])))
        return await self._import_records(records)

    async def import_xlsx(self, content: bytes) -> dict:
        """导入 xlsx（二进制），返回 {success, failures:[{line, errors}]}"""
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            raise BadRequestException("无法解析的 xlsx 文件")
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise BadRequestException("xlsx 缺少数据行")

        header = [_normalize_header(str(h)) if h is not None else "" for h in rows[0]]
        if header != CSV_COLUMNS:
            raise BadRequestException("xlsx 表头不正确，应为: " + ",".join(CSV_COLUMNS))

        records = []
        for row in rows[1:]:
            if not row or not any(v is not None and str(v).strip() for v in row):
                continue
            padded = [str(v).strip() if v is not None else "" for v in row] + [""] * (len(CSV_COLUMNS) - len(row))
            records.append(dict(zip(CSV_COLUMNS, padded[: len(CSV_COLUMNS)])))
        return await self._import_records(records)

    async def _import_records(self, records: list[dict]) -> dict:
        """校验并写入记录（CSV / xlsx 共用），返回 {success, failures}"""
        projects = await self.project_repo.get_active_all()
        code_to_project = {p.code: p for p in projects}

        success = 0
        failures = []
        for idx, record in enumerate(records, start=2):
            errors = []
            project = code_to_project.get(record["项目编码"])
            if not project:
                errors.append(f"项目编码不存在或已停用: {record['项目编码']}")
            if not record["标题"]:
                errors.append("标题不能为空")
            if not record["模块"]:
                errors.append("模块不能为空")
            if not record["预期结果"]:
                errors.append("预期结果不能为空")

            priority = record["优先级"] or "P1"
            if priority not in ALLOWED_PRIORITIES:
                errors.append(f"优先级不合法: {priority}")
            case_type = record["类型"] or "function"
            if case_type not in ALLOWED_CASE_TYPES:
                errors.append(f"类型不合法: {case_type}")
            status = record["状态"] or "draft"
            if status not in ALLOWED_STATUS:
                errors.append(f"状态不合法: {status}")

            module_code = record.get("模块编码", "").strip() or None
            case_code = record.get("用例编码", "").strip() or None

            try:
                validate_codes(module_code, case_code)
                if module_code and case_code:
                    await self._check_uniqueness(project.id, module_code, case_code)
            except BadRequestException as e:
                errors.append(str(e))

            if errors:
                failures.append({"line": idx, "errors": errors})
                continue

            tc = TestCase(
                project_id=project.id,
                title=record["标题"],
                module=record["模块"],
                priority=priority,
                case_type=case_type,
                source=record["来源"] or None,
                precondition=record["前置条件"] or None,
                steps=record["步骤"] or None,
                expected_result=record["预期结果"],
                status=status,
                tags=record["标签"] or None,
                module_code=module_code,
                case_code=case_code,
            )
            self.db.add(tc)
            # 生成文件不影响入库，只记录生成信息
            if project.auto_root_path and module_code and case_code:
                result = await self._generate_auto_file(project, tc)
                if not result["generated"]:
                    # 生成失败但入库成功，记录警告
                    errors.append(result["message"])
            success += 1

        await self.db.flush()
        return {"success": success, "failures": failures}
